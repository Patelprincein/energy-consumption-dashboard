import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from matplotlib.ticker import FuncFormatter

def analyze_and_visualize():
    print("Starting Analysis & Visualization...")
    os.makedirs('dashboard', exist_ok=True)
    conn = sqlite3.connect('energy_data.db')
    
    # Load clean data
    df = pd.read_sql("SELECT * FROM clean_energy_data", conn)
    df['timestamp'] = pd.to_datetime(df['timestamp'])
    
    # -------------------------------------------------------------
    # PHASE 4: TIME-SERIES ANALYSIS
    # -------------------------------------------------------------
    print("Analyzing Time-Series Data...")
    
    # A) Seasonal Patterns (Average Daily Consumption by Month)
    df['month'] = df['timestamp'].dt.month
    monthly_seasonality = df.groupby(['region', 'month'])['consumption_mw'].mean().unstack(level=0)
    
    # B) Demand Peaks (Identify TOP 10 highest usage hours per region)
    peaks_df = df.loc[df.groupby('region')['consumption_mw'].idxmax()]
    
    # Calculate Total Energy Consumed per Region to show efficiencies
    total_consumption = df.groupby('region')['consumption_mw'].sum()

    # -------------------------------------------------------------
    # PHASE 5: VISUALIZATIONS (Matplotlib)
    # -------------------------------------------------------------
    print("Generating Charts...")
    plt.style.use('ggplot') # Optional: makes charts look nicer
    
    # Visualization 1: Seasonal Trends Line Chart
    fig1, ax1 = plt.subplots(figsize=(10, 6))
    monthly_seasonality.plot(ax=ax1, marker='o')
    ax1.set_title('Seasonal Electricity Demand by Region (2024)', fontsize=14, fontweight='bold')
    ax1.set_xlabel('Month', fontsize=12)
    ax1.set_ylabel('Average Demand (MW)', fontsize=12)
    ax1.set_xticks(range(1, 13))
    ax1.set_xticklabels(['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'])
    ax1.legend(title='Region')
    ax1.grid(True, linestyle='--', alpha=0.7)
    plt.tight_layout()
    chart1_path = 'dashboard/seasonal_trends.png'
    fig1.savefig(chart1_path, dpi=300)
    plt.close(fig1)
    
    # Visualization 2: Total Consumption Bar Chart
    fig2, ax2 = plt.subplots(figsize=(8, 6))
    bars = total_consumption.plot(kind='bar', ax=ax2, color=['#1f77b4', '#ff7f0e', '#2ca02c'])
    ax2.set_title('Total Annual Electricity Consumption', fontsize=14, fontweight='bold')
    ax2.set_ylabel('Total Expected MWh', fontsize=12)
    ax2.set_xlabel('Region', fontsize=12)
    
    # format y axis
    ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, p: format(int(x), ',')))
    
    # Annotate bars
    for container in ax2.containers:
        ax2.bar_label(container, fmt='{:,.0f}', padding=3)

    plt.xticks(rotation=0)
    plt.tight_layout()
    chart2_path = 'dashboard/total_consumption.png'
    fig2.savefig(chart2_path, dpi=300)
    plt.close(fig2)

    # -------------------------------------------------------------
    # PHASE 5 (Bonus): Build the Excel Dashboard
    # -------------------------------------------------------------
    print("Building Excel Dashboard...")
    wb = Workbook()
    
    # Sheet 1: Dashboard
    ws_dash = wb.active
    ws_dash.title = "Analytics_Dashboard"
    
    from openpyxl.styles import Font
    
    # Add title and some context text to the Excel Dashboard
    ws_dash['A1'] = "Energy Consumption Analytics Dashboard"
    ws_dash['A1'].font = Font(size=18, bold=True)
    ws_dash['A3'] = "This dashboard highlights seasonal demand patterns and regional consumption totals."
    
    # Insert images
    img1 = OpenpyxlImage(chart1_path)
    img1.width = int(img1.width * 0.6)
    img1.height = int(img1.height * 0.6)
    ws_dash.add_image(img1, 'B5')
    
    img2 = OpenpyxlImage(chart2_path)
    img2.width = int(img2.width * 0.6)
    img2.height = int(img2.height * 0.6)
    ws_dash.add_image(img2, 'J5')
    
    # Sheet 2: Raw Peaks Data Table
    ws_data = wb.create_sheet(title="Demand_Peaks")
    peaks_export = df.sort_values(by='consumption_mw', ascending=False).head(50)
    
    # Write Headers
    headers = list(peaks_export.columns)
    for col_idx, col_name in enumerate(headers, 1):
        ws_data.cell(row=1, column=col_idx, value=col_name)
    
    # Write Data
    for r_idx, row in enumerate(peaks_export.values, 2):
        # Convert timestamp to string so excel handles it nicely if needed, or leave as datetime
        ws_data.cell(row=r_idx, column=1, value=str(row[0])) # timestamp
        ws_data.cell(row=r_idx, column=2, value=row[1]) # region
        ws_data.cell(row=r_idx, column=3, value=row[2]) # mw
        ws_data.cell(row=r_idx, column=4, value=row[3]) # month

    dash_path = 'dashboard/Energy_Dashboard.xlsx'
    wb.save(dash_path)
    print(f"Excel Dashboard saved to: {dash_path}")
    print("Analysis and Visualization completed!")

if __name__ == "__main__":
    analyze_and_visualize()
